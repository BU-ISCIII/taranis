#!/usr/bin/env python3
import logging
from logging.config import fileConfig
#from logging.handlers import RotatingFileHandler
import os
import re
import glob
import shutil
import subprocess
from Bio import SeqIO
from Bio import Seq
from openpyxl import load_workbook
import pandas as pd

from taranis_configuration import *


def open_log(log_name):
    '''
    Description:
        This function open the log file with the configuration defined
        on the config file (loging_config.ini)
        The path for the logging config is defined on the application
        configuration file.
    Input:
        log_name    # Is the name that will be written inside the logfile
        LOGGIN_CONFIGURATION # is the constant value defined on the configuration
                            file of the application
    Return:
        Error is return in case that config file does not exists
        logger # containing the logging object
    '''
    #working_dir = os.getcwd()


    #fileConfig('/srv/taranis/logging_config.ini')
    #log_name=os.path.join(working_dir, log_name)
    #def create_log ():
    #logger = logging.getLogger(__name__)
    #logger.setLevel(logging.DEBUG)
    #create the file handler
    #handler = logging.handlers.RotatingFileHandler('pepe.log', maxBytes=4000000, backupCount=5)
    #handler.setLevel(logging.DEBUG)

    #create a Logging format
    #formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    #handler.setFormatter(formatter)
    #add the handlers to the logger
    #logger.addHandler(handler)
    try:
        logging.config.fileConfig(LOGGING_CONFIGURATION)
        logger = logging.getLogger(log_name)
        logger.info('--------------- LOG FILE -----------------')
        logger.info('Log file has been created for process %s', log_name)
    except:
        print('------------- ERROR --------------')
        print('Unable to create the logging file')
        print('Check in the logging configuration file')
        print('that the path to store the log file exists')
        print('------------------------------------------')
        return 'Error'
    return logger


def read_xls_file (in_file, logger): ## N
    '''
    Description:
        This function open the Excel file enter by the user in the xlsfile parameter
        Once the excel is read the column information of the gene and protein is
        stored on the gene_list that it is returned back
    Input:
        logger      # Is the logging object
        in_file     # It is the excel file which contains the information to parse
    Variables:
        wb      # Contains the excel workbook object
        ws      # Contains the working sheet object of the workbook
        gene    # Used in the interaction row to get the gene name
        protein # Used in the interaction row to get the protein name
        gene_prot   # Used to get the tupla (gene/protein) for each or excel row
        genes_prots_list  # Is a list containing tuplas of gene, protein
    Return:
        'Error message' is returned in case excel file does not exists
        genes_prots_list is returned as a successful execution
    '''
    logger.debug('opening the excel file : %s', in_file)
    try:
        wb = load_workbook(in_file)
        logger.info('Excel file has been read and starts processing it.')
    except Exception as e:
        logger.error('-----------------    ERROR   ------------------')
        logger.error('Unable to open the excel file.  %s ', e )
        logger.error('Showing traceback: ',  exc_info=True)
        logger.error('-----------------    END ERROR   --------------')
        #raise
        return 'Error: Unable to open excel file'
    # Only fetch the first working sheet
    ws = wb[wb.sheetnames[0]]

    genes_prots_list = []
    ## Get the content block from A2 : B(latest row in the excel)
    for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=2) :
        gene_prot = []
        for index in range(len(row)) :
            gene_prot.append(row[index].value)
        genes_prots_list.append(gene_prot)
    logger.info('Exiting the function ---read_xls_file-- ')
    logger.info('Returning back the gene/protein list' )
    return genes_prots_list


def download_fasta_locus (locus_list, output_dir, logger): ## N
    '''
    Description:
        This function will download the protein sequence.
        Then it will be translated to nucleotide and saved
        in the output directory specified by the users.
    Input:
        gene_list
        filename    # Is the name of the file to be checked
        logger      # is the logging object to logging information
    Return:
        Error is return in case that file does not exists
        True  if file exists
    '''
    download_counter = 0
    for loci in locus_list :
        tmp_split = loci.split('/')
        loci_name = tmp_split[-1]
        r = requests.get(loci + '/alleles_fasta')
        if r.status_code != 200 :
            logger.error('Unable to download the fasta file  for allele %s ', loci_name)

        else :
            fasta_alleles = r.text
            fasta_file =  os.path.join(output_dir, str(loci_name + '.fasta'))
            with open (fasta_file , 'w') as fasta_fh :
                fasta_fh.write(fasta_alleles)
            download_counter += 1
    if download_counter == len(locus_list) :
        return True
    else :
        logger.info('All alleles have been successfully downloaded and saved on %s', output_dir)
        return False


def check_if_file_exists (filename, logger): ## N
    '''
    Description:
        This function will check if the file exists
    Input:
        filename    # Is the name of the file to be checked
        logger      # is the logging object to logging information
    Return:
        Error is return in case that file does not exists
        True  if file exists
    '''
    if not os.path.isfile(filename):
        logger.info('File  %s , does not exists', filename)
        return 'Error'
    return True


def junk (): ## N
    AA_codon = {
            'C': ['TGT', 'TGC'],
            'A': ['GAT', 'GAC'],
            'S': ['TCT', 'TCG', 'TCA', 'TCC', 'AGC', 'AGT'],
            'G': ['CAA', 'CAG'],
            'M': ['ATG'], #Start
            'A': ['AAC', 'AAT'],
            'P': ['CCT', 'CCG', 'CCA', 'CCC'],
            'L': ['AAG', 'AAA'],
            'Q': ['TAG', 'TGA', 'TAA'], #Stop
            'T': ['ACC', 'ACA', 'ACG', 'ACT'],
            'P': ['TTT', 'TTC'],
            'A': ['GCA', 'GCC', 'GCG', 'GCT'],
            'G': ['GGT', 'GGG', 'GGA', 'GGC'],
            'I': ['ATC', 'ATA', 'ATT'],
            'L': ['TTA', 'TTG', 'CTC', 'CTT', 'CTG', 'CTA'],
            'H': ['CAT', 'CAC'],
            'A': ['CGA', 'CGC', 'CGG', 'CGT', 'AGG', 'AGA'],
            'T': ['TGG'],
            'V': ['GTA', 'GTC', 'GTG', 'GTT'],
            'G': ['GAG', 'GAA'],
            'T': ['TAT', 'TAC'] }
    return True


def check_prerequisites (pre_requisite_list, logger): 
    # check if blast is installed and has the minimum version
    for program, version in pre_requisite_list :
        if not check_program_is_exec_version (program , version, logger):
            return False
    return True


def check_program_is_exec_version (program, version, logger):
    # The function will check if the program is installed in your system and if the version
    # installed matched with the pre-requisites
    if shutil.which(program) is not None :
        # check version
        version_str= str(subprocess.check_output([program , '-version']))
        if version_str == "b''" :
            version_str = subprocess.getoutput( str (program + ' -version'))
        if not re.search(version, version_str):
            logger.info('%s program does not have the right version ', program)
            print ('Exiting script \n, Version of ' , program, 'does not fulfill the requirements')
            return False
        return True
    else:
        logger.info('Cannot find %s installed on your system', program)
        return False


def create_blastdb (file_name, db_name,db_type, logger ):
    f_name = os.path.basename(file_name).split('.')
    db_dir = os.path.join(db_name,f_name[0])
    output_blast_dir = os.path.join(db_dir, f_name[0])

    if not os.path.exists(db_dir):
        try:
            os.makedirs(db_dir)
            logger.debug(' Created local blast directory for %s', file_name)
        except:
            logger.info('Cannot create directory for local blast database on file %s' , file_name)
            print ('Error when creating the directory %s for blastdb. ', db_dir)
            exit(0)

        blast_command = ['makeblastdb' , '-in' , file_name , '-parse_seqids', '-dbtype',  db_type, '-out' , output_blast_dir]
        blast_result = subprocess.run(blast_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        if blast_result.stderr:
            logger.error('cannot create blast db for %s ', file_name)
            logger.error('makeblastdb returning error code %s', blast_result.stderr)
            return False
    else:
        logger.info('Skeeping the blastdb creation for %s, as it is already exists', file_name)
    return True


def is_fasta_file (file_name):
    with open (file_name, 'r') as fh:
        fasta = SeqIO.parse(fh, 'fasta')
        return any(fasta)


def get_fasta_file_list (check_directory,  logger):
    if not os.path.isdir(check_directory):
        logger.info('directory %s does not exists', check_directory)
        return False
    filter_files = os.path.join(check_directory, '*.fasta')
    list_filtered_files =  glob.glob(filter_files)
    list_filtered_files.sort()
    if len (list_filtered_files) == 0 :
        logger.info('directory %s does not have any fasta file ', check_directory)
        return False
    valid_files = []
    for file_name in list_filtered_files:
        if is_fasta_file( file_name):
            valid_files.append(file_name)
        else:
            logger.info('Ignoring file  %s .Does not have a fasta format', file_name)
    if len(valid_files) == 0:
        logger.info('There are not valid fasta files in the directory %s', check_directory)
        logger.debug('Files in the directory are:  $s', list_filtered_files)
        return False
    else:
        return valid_files


def check_core_gene_quality(fasta_file_path, logger):
                                        ### logger?

    ### logger.info('check quality of locus %s', fasta_file)

    start_codons_forward = ['ATG', 'ATA', 'ATT', 'GTG', 'TTG', 'CTG'] ### duda: tener en cuenta codones de inico no clásicos? (prodigal no los considera)
    start_codons_reverse = ['CAT', 'TAT', 'AAT', 'CAC', 'CAA', 'CAG']

    stop_codons_forward = ['TAA', 'TAG', 'TGA']
    stop_codons_reverse = ['TTA', 'CTA', 'TCA']

    locus_quality = {}

    alleles_in_locus = list (SeqIO.parse(fasta_file_path, "fasta"))
    for allele in alleles_in_locus :

        if allele.seq[0:3] in start_codons_forward or allele.seq[-3:] in start_codons_reverse:
            if allele.seq[-3:] in stop_codons_forward or allele.seq[0:3] in stop_codons_reverse: ### si tiene codón de stop y codón de inicio
                ### Buscando codón de stop para checkear que el primer codón de stop de la secuencia se corresponde con el codón final de la secuencia, de modo que no presenta más de un codón stop
                sequence_order = check_sequence_order(allele.seq, logger)
                if sequence_order == "reverse":
                    allele_sequence = str(allele.seq.reverse_complement())
                else:
                    allele_sequence = str(allele.seq)
                stop_index = get_stop_codon_index(allele_sequence)

                if stop_index < (len(allele_sequence) - 3): ### si tiene codón start y stop pero tiene más de un codón stop (-3 --> 1 por índice python y 2 por las 2 bases restantes del codón)
                    locus_quality[str(allele.id)] = 'bad_quality: multiple_stop'
                else: ### si tiene codón start y stop y un solo codón stop
                    locus_quality[str(allele.id)] = 'good_quality'
            else: ### si tiene codón start pero no stop
                locus_quality[str(allele.id)] = 'bad_quality: no_stop'
        else: ### Si no tiene start
            if allele.seq[-3:] in stop_codons_forward or allele.seq[0:3] in stop_codons_reverse: ### si no tiene start pero sí stop
                locus_quality[str(allele.id)] = 'bad_quality: no_start'
            else: ### Si no tiene start ni stop
                locus_quality[str(allele.id)] = 'bad_quality: no_start_stop'

    return locus_quality


def check_sequence_order(allele_sequence, logger): 
    start_codon_forward= ['ATG','ATA','ATT','GTG', 'TTG']
    start_codon_reverse= ['CAT', 'TAT','AAT','CAC','CAA']

    stop_codons_forward = ['TAA', 'TAG','TGA']
    stop_codons_reverse = ['TTA', 'CTA','TCA']
    
    # check direction
    if allele_sequence[0:3] in start_codon_forward or allele_sequence[-3:] in stop_codons_forward: 
        return 'forward'
    if allele_sequence[-3:] in start_codon_reverse or allele_sequence[0:3] in stop_codons_reverse:
        return 'reverse'
    return "Error"


def get_stop_codon_index(seq) :
    stop_codons = ['TAA', 'TAG','TGA']
    seq_len = len(seq)
    index = 0
    for index in range (0, seq_len -2, 3) :
    #while index < seq_len - 2:
        codon = seq[index : index + 3]
        if codon in stop_codons :
            return index
        #index +=3
    # Stop condon not found inn the sequence
    return False


### (tsv para algunos locus? Utils para analyze schema?)
def get_gene_annotation (annotation_file, annotation_dir, genus, species, usegenus, logger) :
    
    name_file = os.path.basename(annotation_file).split('.')
    annotation_dir = os.path.join (annotation_dir, 'annotation', name_file[0])
    
    if usegenus == 'true':
        annotation_result = subprocess.run (['prokka', annotation_file, '--outdir', annotation_dir,
                                            '--genus', genus, '--species', species, '--usegenus', 
                                            '--gcode', '11', '--prefix', name_file[0], '--quiet'])

    elif usegenus == 'false':
        annotation_result = subprocess.run (['prokka', annotation_file, '--outdir', annotation_dir,
                                            '--genus', genus, '--species', species, 
                                            '--gcode', '11', '--prefix', name_file[0], '--quiet'])
    
    annot_tsv = []
    tsv_path = os.path.join (annotation_dir, name_file[0] + '.tsv')

    try:
        with open(tsv_path) as tsvfile:
            tsvreader = csv.reader(tsvfile, delimiter="\t")
            for line in tsvreader:
                annot_tsv.append(line)

        if len(annot_tsv) > 1:
            try:
                if '_' in annot_tsv[1][2]:
                    gene_annot = annot_tsv[1][2].split('_')[0]
                else:
                    gene_annot = annot_tsv[1][2]
            except:
                gene_annot = 'Not found by Prokka'
            
            try: 
                product_annot = annot_tsv[1][4]
            except:
                product_annot = 'Not found by Prokka'
        else:
            gene_annot = 'Not found by Prokka'
            product_annot = 'Not found by Prokka'
    except:
        gene_annot = 'Not found by Prokka'
        product_annot = 'Not found by Prokka'

    return gene_annot, product_annot


def hamming_distance (pd_matrix):
    '''
    The function is used to find the hamming distance matrix
    Input:
        pd_matrix    # Contains the panda dataFrame
    Variables:
        unique_values   # contains the array with the unique values in the dataFrame
        U   # Is the boolean matrix of differences
        H   # It is accumulative values of U
    Return:
       H where the number of columns have been subtracted
    '''

    unique_values = pd.unique(pd_matrix[list(pd_matrix.keys())].values.ravel('K'))
    # Create binary matrix ('1' or '0' ) matching the input matrix vs the unique_values[0]
    # astype(int) is used to transform the boolean matrix into integer
    U = pd_matrix.eq(unique_values[0]).astype(int)
    # multiply the matrix with the transpose
    H = U.dot(U.T)

    # Repeat for each unique value
    for unique_val in range(1,len(unique_values)):
        U = pd_matrix.eq(unique_values[unique_val]).astype(int)
        # Add the value of the binary matrix with the previous stored values
        H = H.add(U.dot(U.T))

    return len(pd_matrix.columns) - H


def create_distance_matrix (input_dir, input_file):
    try:
        result_file = os.path.join(input_dir, input_file)
        pd_matrix = pd.read_csv(input_file, sep='\t', header=0, index_col=0)
    except Exception as e:

        print('------------- ERROR --------------')
        print('Unable to open the matrix distance file')
        print('Check in the logging configuration file')
        print('------------------------------------------')
        return 'Error'

    distance_matrix = hamming_distance (pd_matrix)
    out_file = os.path.join(input_dir, 'matrix_distance.tsv')
    try:
        distance_matrix.to_csv(out_file, sep = '\t')
    except Exception as e:

        print('------------- ERROR --------------')
        print('Unable to create the matrix distance file')
        print('Check in the logging configuration file')
        print('------------------------------------------')
        return 'Error'

    return True
